from tkinter import messagebox, ttk
from datetime import date, datetime, timedelta
from tkcalendar import DateEntry
from dotenv import load_dotenv
import tkinter as tk
import pandas as pd
import threading
import requests
import os


# --- CONFIGURAÇÕES ---
DRIVE_PATH = r"\\100.64.1.10\Users\pichau\Documents\Drive Comagro"

# Carrega variáveis de ambiente do caminho de rede especificado
# ENV_PATH = r"\\100.64.1.10\Ti Compartilhado\Financeiro\.env"
ENV_PATH = os.path.join(DRIVE_PATH, "Ti Compartilhado", "Financeiro", ".env")
load_dotenv(ENV_PATH)

TEMP_DIR = os.path.join(os.getcwd(), "temp_xfin")
CONFIG_FILE = os.path.join(DRIVE_PATH, "CONTAS A PAGAR", "0 Extras", "[CONFIG] Dados Bancários Fornecedores.xlsx")
XFIN_URL = "https://app.xfin.com.br"
TK_XFIN = os.getenv('TK_XFIN')

# Configurações do Banco de Dados (Firebird)
FB_HOST = os.getenv('HOST')
FB_PORT = int(os.getenv('PORT', '3050'))
FB_DB = os.getenv('DB_PATH')
FB_USER = os.getenv('APP_USER')
FB_PASS = os.getenv('PASSWORD')
FB_ROLE = os.getenv('ROLE')
FB_AUTH = os.getenv('AUTH')

# Mapeamento de Colunas do Xfin (Resposta da API)
COL_XFIN_FORNECEDOR = "pessoa"
COL_XFIN_VENCIMENTO = "dataVencimento"
COL_XFIN_VALOR = "valor"
COL_XFIN_DOC = "numeroDocumento"
COL_XFIN_OBS = "descricao"
COL_XFIN_FORMA_PAGTO = "tipoDocumento"
COL_XFIN_BANCO_PAGAR = "banco"
COL_XFIN_FILIAL = "filial"

# --- FUNÇÕES AUXILIARES ---


def get_firebird_connection():
    import firebirdsql
    try:
        return firebirdsql.connect(
            host=FB_HOST,
            port=FB_PORT,
            database=FB_DB,
            user=FB_USER,
            password=FB_PASS,
            role=FB_ROLE,
            auth_plugin_name=FB_AUTH,
            wire_crypt=False,
            charset='ISO8859_1'
        )
    except Exception as e:
        print(f"Erro ao conectar ao Firebird: {e}")
        return None


def check_drive_access():
    if not os.path.exists(DRIVE_PATH):
        raise Exception(f"Drive de rede inacessível: {DRIVE_PATH}")
    return True


def get_date_range():
    """
    Calcula o intervalo de datas para o arquivo de saída.
    Se hoje for Sábado -> Sábado, Domingo, Segunda.
    Se hoje for Domingo -> Domingo, Segunda.
    Caso contrário -> Hoje.
    """
    today = date.today()
    weekday = today.weekday()  # 0=Seg, 5=Sab, 6=Dom

    start_date = today
    end_date = today

    if weekday == 5:  # Sábado
        end_date = today + timedelta(days=2)  # Até Segunda
    elif weekday == 6:  # Domingo
        end_date = today + timedelta(days=1)  # Até Segunda

    return start_date, end_date


def format_currency(value):
    if pd.isna(value):
        return "R$ 0,00"
    try:
        return f"R$ {float(value):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except:
        return str(value)


def create_default_config(path):
    """Cria o arquivo de configuração padrão com as colunas necessárias."""
    try:
        # Garante que o diretório existe
        os.makedirs(os.path.dirname(path), exist_ok=True)

        # Colunas necessárias para o funcionamento do robô
        columns = [
            "Fornecedor",
            "CNPJ",
            "Chave PIX",
            "Banco",
            "Nome Titular"
        ]

        df = pd.DataFrame(columns=columns)
        df.to_excel(path, index=False)
        print(f"Arquivo de configuração criado em: {path}")
    except Exception as e:
        raise Exception(f"Erro ao criar arquivo de configuração padrão: {e}")


def identify_branch_group(val):
    """Identifica o grupo da filial baseado no CNPJ ou Nome."""
    val = str(val).upper()

    # 1. Verificação por CNPJ (Mais preciso)
    if "62.188.494" in val:
        return "Servicos"
    if "59.185.879" in val:
        return "Divisa"
    if "14.255.350" in val:
        return "Comagro"  # Loja (0001) e Oficina (0004)

    # 2. Verificação por Nome
    if "DIVISA" in val:
        return "Divisa"

    # Verificar PEÇAS ou OFICINA antes de SERVIÇOS para evitar ambiguidade
    if "PEÇAS" in val or "PECAS" in val or "OFICINA" in val:
        return "Comagro"

    if "SERVI" in val and "COMAGRO" in val:
        return "Servicos"
    if "COMAGRO" in val:
        return "Comagro"

    return "Geral"


def identify_filial_tab(val):
    """
    Identifica a filial para aba e ordenação do Excel.
    Loja=1, Oficina=2, Divisa=3, Serviços=4.
    Retorna tupla (ordem, nome_aba).
    """
    val = str(val).upper()

    # 1. Verificação por CNPJ (mais preciso)
    if "62.188.494" in val:
        return (4, "PAGAMENTOS - SERVICOS")
    if "59.185.879" in val:
        return (3, "PAGAMENTOS - DIVISA")
    if "0004" in val:
        return (2, "PAGAMENTOS - OFICINA")
    if "0001" in val or "14.255.350" in val:
        return (1, "PAGAMENTOS - LOJA")

    # 2. Verificação por Nome
    if "DIVISA" in val:
        return (3, "PAGAMENTOS - DIVISA")
    # OFICINA antes de PEÇAS para evitar ambiguidade
    if "OFICINA" in val:
        return (2, "PAGAMENTOS - OFICINA")
    # PEÇAS antes de SERVIÇOS — "Comagro Peças e Serviços" é a razão social da loja, não a filial de Serviços
    if "PEÇAS" in val or "PECAS" in val or "LOJA" in val:
        return (1, "PAGAMENTOS - LOJA")
    if "SERVI" in val and "COMAGRO" in val:
        return (4, "PAGAMENTOS - SERVICOS")
    if "COMAGRO" in val:
        return (1, "PAGAMENTOS - LOJA")

    return (99, "GERAL")


def get_file_date(dt):
    """Agrupa datas de fim de semana para a segunda-feira."""
    if pd.isna(dt):
        return date.today()
    if isinstance(dt, pd.Timestamp):
        dt = dt.date()

    weekday = dt.weekday()
    if weekday == 5:  # Sábado -> Segunda
        return dt + timedelta(days=2)
    if weekday == 6:  # Domingo -> Segunda
        return dt + timedelta(days=1)
    return dt

# --- INTEGRAÇÃO API XFIN ---


def fetch_xfin_data_api(status_callback, dt_ini, dt_fim, stop_event):

    if not TK_XFIN:
        raise Exception("Token XFIN (TK_XFIN) não encontrado no arquivo .env. Configure-o para acessar a API.")

    status_callback("Iniciando busca na API...")

    # A API exige o formato yyyy-MM-dd
    dt_ini_api = datetime.strptime(dt_ini, "%d/%m/%Y").strftime("%Y-%m-%d")
    dt_fim_api = datetime.strptime(dt_fim, "%d/%m/%Y").strftime("%Y-%m-%d")

    url = f"{XFIN_URL}/api/v1/contasPagar"
    headers = {
        "accept": "*/*",
        "Authorization": f"Bearer {TK_XFIN}"
    }

    tamanho_pagina = 500
    pagina = 1
    todos_itens = []

    while True:
        if stop_event.is_set():
            return pd.DataFrame()

        status_callback(f"Buscando dados na API... Página {pagina}")
        params = {
            "pagina": pagina,
            "tamanhoPagina": tamanho_pagina,
            "dataVencimentoInicial": dt_ini_api,
            "dataVencimentoFinal": dt_fim_api
        }

        try:
            response = requests.get(url, headers=headers, params=params)
            response.raise_for_status()
            data = response.json()

            if not data.get("sucesso", False):
                raise Exception(f"Erro na API: {data.get('mensagem', 'Desconhecido')}")

            itens = data.get("itens", [])
            todos_itens.extend(itens)

            total_paginas = data.get("totalPaginas", 1)
            if pagina >= total_paginas:
                break

            pagina += 1

        except Exception as e:
            print(f"Erro ao acessar API na página {pagina}: {e}")
            raise Exception(f"Falha na comunicação com a API XFIN: {e}")

    return pd.DataFrame(todos_itens)

# --- PROCESSAMENTO DE DADOS ---


def process_data(df_xfin, status_callback, stop_event, dt_ini_ui=None, dt_fim_ui=None):
    import re
    status_callback("Processando dados da API...")

    if df_xfin is None or df_xfin.empty:
        return None, [], date.today(), date.today(), None

    col_fornecedor = COL_XFIN_FORNECEDOR
    col_vencimento = COL_XFIN_VENCIMENTO
    col_valor = COL_XFIN_VALOR
    col_doc = COL_XFIN_DOC
    col_obs = COL_XFIN_OBS
    col_forma = COL_XFIN_FORMA_PAGTO
    col_banco = COL_XFIN_BANCO_PAGAR
    col_filial = COL_XFIN_FILIAL

    def _parse_vencimento_series(series: pd.Series, dt_ini_str: str | None, dt_fim_str: str | None) -> pd.Series:
        """
        Converte a série de datas do Xfin para datetime de forma robusta.

        Problema que isso resolve:
        - Alguns payloads podem vir como '04/01/2026' (MM/DD/YYYY) e outros como '01/04/2026' (DD/MM/YYYY).
        - Se interpretarmos errado, '01/04' (1º de Abril) vira 4 de Janeiro e o robô salva em pastas "trocadas".
        """

        # Intervalo escolhido na UI (sempre dd/mm/yyyy), usado para desambiguar datas com barras
        ini = None
        fim = None
        try:
            if dt_ini_str:
                ini = datetime.strptime(dt_ini_str, "%d/%m/%Y").date()
            if dt_fim_str:
                fim = datetime.strptime(dt_fim_str, "%d/%m/%Y").date()
        except Exception:
            ini = None
            fim = None

        s = series.copy()
        # Se já for datetime, mantém
        if pd.api.types.is_datetime64_any_dtype(s):
            return s

        s_str = s.astype(str).str.strip()
        s_str = s_str.replace({"None": "", "nan": "", "NaT": ""})

        out = pd.Series(pd.NaT, index=s_str.index, dtype="datetime64[ns]")

        # 1) ISO (yyyy-mm-dd...) é não-ambíguo
        iso_mask = s_str.str.match(r"^\d{4}-\d{2}-\d{2}")
        if iso_mask.any():
            out.loc[iso_mask] = pd.to_datetime(s_str.loc[iso_mask], errors="coerce")

        # 2) Datas com barras (dd/mm/yyyy ou mm/dd/yyyy)
        slash_mask = s_str.str.match(r"^\d{1,2}/\d{1,2}/\d{2,4}$")
        if slash_mask.any():
            slash_vals = s_str.loc[slash_mask]
            parts = slash_vals.str.split("/", expand=True)
            p1 = pd.to_numeric(parts[0], errors="coerce")
            p2 = pd.to_numeric(parts[1], errors="coerce")

            # Casos não-ambíguos
            ddmm_mask = (p1 > 12)  # 13/04/... só pode ser dd/mm
            mmdd_mask = (p2 > 12)  # 04/13/... só pode ser mm/dd

            if ddmm_mask.any():
                idx = slash_vals.index[ddmm_mask]
                out.loc[idx] = pd.to_datetime(slash_vals.loc[idx], dayfirst=True, errors="coerce")
            if mmdd_mask.any():
                idx = slash_vals.index[mmdd_mask]
                out.loc[idx] = pd.to_datetime(slash_vals.loc[idx], dayfirst=False, errors="coerce")

            # Casos ambíguos (1..12 / 1..12)
            amb_mask = ~(ddmm_mask | mmdd_mask)
            if amb_mask.any():
                idx = slash_vals.index[amb_mask]
                parsed_dayfirst = pd.to_datetime(slash_vals.loc[idx], dayfirst=True, errors="coerce")
                parsed_monthfirst = pd.to_datetime(slash_vals.loc[idx], dayfirst=False, errors="coerce")

                # Se tivermos intervalo da UI, escolhe a interpretação que mais cai dentro do intervalo
                if ini is not None and fim is not None:
                    d1 = parsed_dayfirst.dt.date
                    d2 = parsed_monthfirst.dt.date
                    score1 = ((d1 >= ini) & (d1 <= fim)).sum()
                    score2 = ((d2 >= ini) & (d2 <= fim)).sum()
                    chosen = parsed_dayfirst if score1 >= score2 else parsed_monthfirst
                else:
                    # Padrão BR quando não dá para desambiguar
                    chosen = parsed_dayfirst

                out.loc[idx] = chosen

        # 3) Fallback: tenta parse genérico
        remaining = out.isna() & (s_str != "")
        if remaining.any():
            out.loc[remaining] = pd.to_datetime(s_str.loc[remaining], errors="coerce")

        return out

    # Converter vencimento para datetime (robusto para DD/MM vs MM/DD)
    # dt_ini_ui/dt_fim_ui vêm da UI e ajudam a desambiguar datas com barras.
    df_xfin[col_vencimento] = _parse_vencimento_series(df_xfin[col_vencimento], dt_ini_ui, dt_fim_ui)

    df_filtered = df_xfin.copy()

    # Determinar range de datas baseado nos dados para nomear o arquivo
    if not df_filtered.empty:
        start_date = df_filtered[col_vencimento].min().date()
        end_date = df_filtered[col_vencimento].max().date()
    else:
        start_date = date.today()
        end_date = date.today()

    if df_filtered.empty:
        return None, [], start_date, end_date, None

    # 2. Ler Configuração Bancária
    status_callback("Lendo dados bancários...")
    if not os.path.exists(CONFIG_FILE):
        status_callback("Criando arquivo de configuração padrão...")
        create_default_config(CONFIG_FILE)

    df_config = pd.read_excel(CONFIG_FILE, dtype=str)
    # Renomear colunas do config para evitar colisão com o CSV do Xfin
    df_config.columns = [f"Config_{c}" if c != "Fornecedor" else c for c in df_config.columns]

    # Normalizar nomes para merge (uppercase, strip)
    df_config['Fornecedor_Norm'] = df_config['Fornecedor'].str.upper().str.strip()

    def clean_supplier_name(name):
        if pd.isna(name):
            return ""
        name = str(name).upper().strip()
        # Remove "[CODE] - " prefix using regex to handle hyphens in name correctly
        match = re.match(r'^(\d+)\s*-\s*(.*)', name)
        if match:
            return match.group(2).strip()
        return name

    df_filtered['Fornecedor_Norm'] = df_filtered[col_fornecedor].apply(clean_supplier_name)

    # 3. Buscar CNPJ no Firebird (Enriquecimento)
    status_callback("Consultando Firebird...")
    conn_fb = get_firebird_connection()
    if stop_event.is_set():
        if conn_fb:
            conn_fb.close()
        return None, [], start_date, end_date, None

    fb_data = {}
    if conn_fb:
        cursor = conn_fb.cursor()
        # Busca todos fornecedores para criar um dict de lookup
        cursor.execute("SELECT NOME, CPF_CNPJ FROM FORNECEDOR")
        for row in cursor.fetchall():
            nome = row[0].strip().upper() if row[0] else ""
            cnpj = row[1].strip() if row[1] else ""
            fb_data[nome] = cnpj
        print(f"Carregados {len(fb_data)} fornecedores do Firebird.")
        conn_fb.close()

    # Aplicar CNPJ do Firebird no DataFrame
    df_filtered['CNPJ_FB'] = df_filtered['Fornecedor_Norm'].map(fb_data)

    # 4. Merge com Configuração Bancária
    status_callback("Cruzando dados...")
    df_merged = pd.merge(df_filtered, df_config, on='Fornecedor_Norm', how='left')

    # Identificar faltantes
    missing_suppliers = df_merged[df_merged['Config_Chave PIX'].isna(
    ) & df_merged['Config_Banco'].isna()]['Fornecedor_Norm'].unique()

    # Preparar dados para Excel
    # A API já retorna valores numéricos, mas garantimos o tipo e tratamos vazios
    df_merged[col_valor] = pd.to_numeric(df_merged[col_valor], errors='coerce').fillna(0.0)

    # Identificar filial para abas do Excel (ordem: Loja, Oficina, Divisa, Serviços)
    if col_filial:
        filial_info = df_merged[col_filial].apply(identify_filial_tab)
        df_merged['Filial_Order'] = filial_info.apply(lambda x: x[0])
        df_merged['Filial_Sheet'] = filial_info.apply(lambda x: x[1])
        # Mantém Filial_Group para compatibilidade com alertas e logs
        df_merged['Filial_Group'] = df_merged[col_filial].apply(identify_branch_group)
    else:
        df_merged['Filial_Order'] = 99
        df_merged['Filial_Sheet'] = 'GERAL'
        df_merged['Filial_Group'] = 'Geral'

    # 5. Definir CNPJ Final (Prioridade: Excel > Firebird)
    if 'Config_CNPJ' in df_merged.columns:
        df_merged['CNPJ_Final'] = df_merged['Config_CNPJ'].fillna(df_merged['CNPJ_FB'])
    else:
        df_merged['CNPJ_Final'] = df_merged['CNPJ_FB']

    # 6. Extrair Fatura da Descrição (Feature Nova - Strict)
    def extract_invoice_strict(row):
        desc = str(row[col_obs]) if col_obs and pd.notna(row[col_obs]) else ""
        if " - " in desc:
            # Pega a última parte após o último hífen
            candidate = desc.rsplit(" - ", 1)[-1].strip()
            # Validação estrita: aceita apenas se for numérico puro (sem /, letras, pontos)
            if candidate.isdigit():
                return candidate
        return ""

    df_merged['Fatura'] = df_merged.apply(extract_invoice_strict, axis=1)

    return df_merged, missing_suppliers, start_date, end_date, (
        col_fornecedor, col_vencimento, col_valor, col_doc, col_obs, col_forma, col_banco)


def clean_sheet_name(name):
    """Remove caracteres inválidos para nome de aba do Excel."""
    invalid_chars = ['\\', '/', '*', '[', ']', ':', '?']
    for char in invalid_chars:
        name = name.replace(char, '')
    # Excel limita a 31 caracteres
    return name[:31]

# --- GERAÇÃO DE EXCEL ---

# Espaçamento entre tabelas empilhadas na mesma aba
GAP_ENTRE_TABELAS = 3
# Tabela de totais por forma de pagamento começa em K3
COL_RESUMO_TIPO = 11  # K
COL_RESUMO_VALOR = 12  # L
LINHA_RESUMO_CABECALHO = 3


def _doc_priority(x):
    """Ordena formas de pagamento: NF, DDA, Boleto, PIX e demais."""
    xu = x.upper()
    if "NF" in xu or "NOTA" in xu:
        return (0, xu)
    if "BOLETO" in xu:
        return (1, xu)
    if "PIX" in xu:
        return (2, xu)
    return (3, xu)


def _get_doc_style(doc_type):
    """Retorna cores de cabeçalho conforme o tipo de pagamento."""
    dt_upper = doc_type.upper()
    if "NF" in dt_upper or "NOTA" in dt_upper:
        return "C4BD96", "000000"  # marrom-claro (0.77, 0.74, 0.59)
    if "BOLETO" in dt_upper:
        return "B2A1C7", "000000"  # lilás (0.70, 0.63, 0.78)
    if "CRÉDITO" in dt_upper or "CREDITO" in dt_upper or "ESTORNO" in dt_upper:
        return "366191", "FFFFFF"  # azul-escuro (0.21, 0.38, 0.57)
    if "PIX" in dt_upper:
        return "FABF8F", "000000"  # pêssego (0.98, 0.75, 0.56)
    return "BFBFBF", "000000"  # cinza (0.75, 0.75, 0.75) — demais tipos


def _prepare_group_df(group_df, is_pix_layout, col_forn, col_valor, col_obs, col_venc):
    """Ordena e agrupa faturas/PIX antes de escrever a tabela."""
    group_df = group_df.copy()

    if is_pix_layout:
        group_df['__Total_Supplier'] = group_df.groupby(col_forn)[col_valor].transform('sum')
        return group_df.sort_values(
            by=['__Total_Supplier', col_forn, col_valor],
            ascending=[True, True, True])

    if 'Fatura' in group_df.columns:
        group_df['Fatura'] = group_df['Fatura'].fillna('').astype(str).str.strip()
        mask_invoice = group_df['Fatura'] != ''
        df_invoice = group_df[mask_invoice].copy()
        df_no_invoice = group_df[~mask_invoice].copy()

        if not df_invoice.empty:
            grp_keys = ['Fatura', col_forn, col_venc]
            if 'CNPJ_Final' in df_invoice.columns:
                df_invoice['CNPJ_Final'] = df_invoice['CNPJ_Final'].fillna('')
                grp_keys.append('CNPJ_Final')
            if 'Config_Banco' in df_invoice.columns:
                df_invoice['Config_Banco'] = df_invoice['Config_Banco'].fillna('')
                grp_keys.append('Config_Banco')

            grouped_rows = []
            for key, block in df_invoice.groupby(grp_keys):
                row_data = block.iloc[0].copy()
                row_data[col_valor] = block[col_valor].sum()
                if col_obs:
                    row_data[col_obs] = f"Fatura - {key[0]}"
                grouped_rows.append(row_data)
            group_df = pd.concat([df_no_invoice, pd.DataFrame(grouped_rows)], ignore_index=True)

    return group_df.sort_values(by=[col_valor], ascending=True)


def _add_schedule_dropdown(ws, row, col, dv_agendamento):
    """Adiciona seletor de agendamento com cor verde/vermelha."""
    from openpyxl.styles import Font, PatternFill
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.utils import get_column_letter

    col_letter = get_column_letter(col)
    cell_ref = f"${col_letter}${row}"

    schedule_cell = ws.cell(row=row, column=col, value="Não Agendado")
    schedule_cell.font = Font(bold=True, color="FFFFFF")
    schedule_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    dv_agendamento.add(schedule_cell)

    green_rule = FormulaRule(
        formula=[f'{cell_ref}="Agendado"'],
        fill=PatternFill(start_color="00B050", end_color="00B050", fill_type="solid"),
        font=Font(bold=True, color="FFFFFF"),
    )
    red_rule = FormulaRule(
        formula=[f'{cell_ref}="Não Agendado"'],
        fill=PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
        font=Font(bold=True, color="FFFFFF"),
    )
    ws.conditional_formatting.add(f"{col_letter}{row}", green_rule)
    ws.conditional_formatting.add(f"{col_letter}{row}", red_rule)


def _write_payment_table(ws, group_df, table_title, doc_type, start_row, cols_map, border, currency_fmt, dv_agendamento):
    """
    Escreve uma tabela de pagamento empilhada na aba.
    Retorna (próxima_linha_livre, referência_célula_total).
    """
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    col_forn, col_venc, col_valor, col_doc, col_obs, col_forma, col_banco = cols_map
    current_row = start_row

    # Título da tabela (nome da forma de pagamento)
    fill_color, font_color = _get_doc_style(doc_type)
    title_cell = ws.cell(row=current_row, column=1, value=table_title.upper())
    title_cell.font = Font(bold=True, color=font_color, size=12)
    title_cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    current_row += 1

    is_pix_layout = "PIX" in doc_type.upper()
    if is_pix_layout:
        headers = ["Vencimento", "Nome Recebedor", "Fornecedor",
                   "Chave PIX", "Nº Doc", "Observação", "Valor", "Valor Total"]
        val_col_idx = 8
    else:
        headers = ["Vencimento", "Banco/Conta", "Fornecedor", "CNPJ", "Nº Doc", "Observação", "Valor"]
        val_col_idx = 7

    header_font = Font(bold=True, color=font_color)
    header_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

    header_row = current_row
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        ws.column_dimensions[get_column_letter(col_num)].width = 20

    group_df = _prepare_group_df(group_df, is_pix_layout, col_forn, col_valor, col_obs, col_venc)
    current_row = header_row + 1
    data_start_row = current_row

    start_merge_row = current_row
    current_supplier = None
    supplier_total = 0.0

    for _, row in group_df.iterrows():
        val = row[col_valor]

        if is_pix_layout:
            supplier = row[col_forn]
            if supplier != current_supplier:
                if current_supplier is not None:
                    if start_merge_row < current_row - 1:
                        ws.merge_cells(start_row=start_merge_row, start_column=8,
                                       end_row=current_row - 1, end_column=8)
                    ws.cell(row=start_merge_row, column=8, value=supplier_total).number_format = currency_fmt
                    ws.cell(row=start_merge_row, column=8).alignment = Alignment(vertical='center')
                current_supplier = supplier
                start_merge_row = current_row
                supplier_total = 0.0
            supplier_total += val

            ws.cell(row=current_row, column=1, value=row[col_venc].strftime('%d/%m/%Y'))
            ws.cell(row=current_row, column=2, value=row.get('Config_Nome Titular', ''))
            ws.cell(row=current_row, column=3, value=supplier)
            ws.cell(row=current_row, column=4, value=row.get('Config_Chave PIX', ''))
            ws.cell(row=current_row, column=5, value=row[col_doc] if col_doc and col_doc in row else "")
            ws.cell(row=current_row, column=6, value=row[col_obs] if col_obs and col_obs in row else "")
            ws.cell(row=current_row, column=7, value=val).number_format = currency_fmt
        else:
            banco_val = row.get('Config_Banco', '')
            ws.cell(row=current_row, column=1, value=row[col_venc].strftime('%d/%m/%Y'))
            ws.cell(row=current_row, column=2, value=banco_val)
            ws.cell(row=current_row, column=3, value=row[col_forn])
            ws.cell(row=current_row, column=4, value=row.get('CNPJ_Final', ''))
            ws.cell(row=current_row, column=5, value=row[col_doc] if col_doc and col_doc in row else "")
            ws.cell(row=current_row, column=6, value=row[col_obs] if col_obs and col_obs in row else "")
            ws.cell(row=current_row, column=7, value=val).number_format = currency_fmt

        ws.column_dimensions['E'].width = 7
        current_row += 1

    if is_pix_layout and current_supplier is not None:
        if start_merge_row < current_row - 1:
            ws.merge_cells(start_row=start_merge_row, start_column=8, end_row=current_row - 1, end_column=8)
        ws.cell(row=start_merge_row, column=8, value=supplier_total).number_format = currency_fmt
        ws.cell(row=start_merge_row, column=8).alignment = Alignment(vertical='center')

    data_end_row = current_row - 1
    total_row = current_row + 1
    col_letter = get_column_letter(val_col_idx)

    ws.cell(row=total_row, column=val_col_idx - 1, value="TOTAL:").font = Font(bold=True)
    if data_end_row >= data_start_row:
        sum_formula = f"=SUM({col_letter}{data_start_row}:{col_letter}{data_end_row})"
    else:
        sum_formula = 0
    c_total = ws.cell(row=total_row, column=val_col_idx, value=sum_formula)
    c_total.number_format = currency_fmt
    c_total.font = Font(bold=True)

    schedule_row = total_row + 2
    _add_schedule_dropdown(ws, schedule_row, 1, dv_agendamento)

    next_row = schedule_row + 1 + GAP_ENTRE_TABELAS
    return next_row, f"{col_letter}{total_row}"


def _write_summary_table(ws, summary_data, border, currency_fmt):
    """Escreve tabela de totais na mesma aba, a partir de K3."""
    from openpyxl.styles import Font, PatternFill

    header_font_tot = Font(bold=True, color="FFFFFF")
    header_fill_tot = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

    ws.column_dimensions['K'].width = 30
    ws.column_dimensions['L'].width = 20

    cell_h1 = ws.cell(row=LINHA_RESUMO_CABECALHO, column=COL_RESUMO_TIPO, value="Tipo de Pagamento")
    cell_h2 = ws.cell(row=LINHA_RESUMO_CABECALHO, column=COL_RESUMO_VALOR, value="Valor Total")
    for c in [cell_h1, cell_h2]:
        c.font = header_font_tot
        c.fill = header_fill_tot
        c.border = border

    r = LINHA_RESUMO_CABECALHO + 1
    first_data_row = r
    for name, cell_ref in summary_data:
        ws.cell(row=r, column=COL_RESUMO_TIPO, value=name)
        c_val = ws.cell(row=r, column=COL_RESUMO_VALOR, value=f"={cell_ref}")
        c_val.number_format = currency_fmt
        r += 1

    if summary_data:
        r += 1
        cell_gt_lbl = ws.cell(row=r, column=COL_RESUMO_TIPO, value="TOTAL GERAL")
        cell_gt_val = ws.cell(row=r, column=COL_RESUMO_VALOR, value=f"=SUM(L{first_data_row}:L{r - 1})")
        for c in [cell_gt_lbl, cell_gt_val]:
            c.font = Font(bold=True, size=12)
            c.border = border
        cell_gt_val.number_format = currency_fmt


def _build_payment_subgroups(df_filial, col_forma, col_banco):
    """Monta subgrupos por forma de pagamento (separando bancos quando necessário)."""
    doc_vals = list(df_filial[col_forma].astype(str).fillna('').unique())
    doc_types = sorted(doc_vals, key=_doc_priority)
    sub_groups = []

    for doc_type in doc_types:
        df_doc = df_filial[df_filial[col_forma] == doc_type].copy()
        unique_banks = df_doc[col_banco].unique()
        real_banks = [b for b in unique_banks if str(b).strip()]

        if len(real_banks) > 1:
            for bank in unique_banks:
                sub_df = df_doc[df_doc[col_banco] == bank]
                if sub_df.empty:
                    continue
                s_name = f"{doc_type}"
                if str(bank).strip():
                    s_name += f" - {bank}"
                sub_groups.append((s_name, doc_type, sub_df))
        else:
            sub_groups.append((doc_type, doc_type, df_doc))

    return sub_groups


def create_excel(df, output_path, cols_map):
    import openpyxl
    from openpyxl.styles import Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation

    col_forn, col_venc, col_valor, col_doc, col_obs, col_forma, col_banco = cols_map

    wb = openpyxl.Workbook()
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # Garantir colunas de agrupamento
    if col_forma and col_forma in df.columns:
        df[col_forma] = df[col_forma].fillna('Indefinido')
    else:
        df['__Forma_Temp'] = 'Indefinido'
        col_forma = '__Forma_Temp'

    if col_banco and col_banco in df.columns:
        df[col_banco] = df[col_banco].fillna('')
    else:
        df['__Banco_Temp'] = ''
        col_banco = '__Banco_Temp'

    currency_fmt = '_-R$* #,##0.00_-;-R$* #,##0.00_-;_-R$* \"-\"??_-;_-@_-'
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    # Uma aba por filial, na ordem Loja -> Oficina -> Divisa -> Serviços
    filiais = (
        df[['Filial_Order', 'Filial_Sheet']]
        .drop_duplicates()
        .sort_values('Filial_Order')
    )

    for _, filial_row in filiais.iterrows():
        sheet_name = clean_sheet_name(filial_row['Filial_Sheet'])
        df_filial = df[df['Filial_Sheet'] == filial_row['Filial_Sheet']].copy()
        if df_filial.empty:
            continue

        ws = wb.create_sheet(sheet_name)

        # Validação de agendamento reutilizada em todas as tabelas da aba
        dv_agendamento = DataValidation(
            type="list",
            formula1='"Agendado,Não Agendado"',
            allow_blank=False,
        )
        dv_agendamento.error = "Selecione Agendado ou Não Agendado"
        dv_agendamento.errorTitle = "Agendamento"
        ws.add_data_validation(dv_agendamento)

        current_row = 1
        sheet_summary = []

        for table_title, doc_type, group_df in _build_payment_subgroups(df_filial, col_forma, col_banco):
            current_row, total_ref = _write_payment_table(
                ws, group_df, table_title, doc_type, current_row,
                cols_map, border, currency_fmt, dv_agendamento,
            )
            sheet_summary.append((table_title, total_ref))

        if sheet_summary:
            _write_summary_table(ws, sheet_summary, border, currency_fmt)

    if not wb.sheetnames:
        wb.create_sheet("VAZIO")

    try:
        wb.save(output_path)
    except PermissionError:
        output_path = output_path + ".error"

    wb.save(output_path)

# --- CLASSE PRINCIPAL DA UI ---


class PaymentBotApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Robô de Pagamentos Xfin")
        self.root.geometry("450x400")

        self.stop_event = threading.Event()

        self.lbl_status = tk.Label(root, text="Pronto para iniciar", wraplength=350)
        self.lbl_status.pack(pady=20)

        # Inputs de Data
        frame_dates = tk.Frame(root)
        frame_dates.pack(pady=5)
        tk.Label(frame_dates, text="Início:").pack(side=tk.LEFT)
        self.entry_start = DateEntry(frame_dates, width=12, background='darkblue',
                                     foreground='white', borderwidth=2, locale='pt_BR', date_pattern='dd/mm/yyyy')
        self.entry_start.pack(side=tk.LEFT, padx=5)
        self.entry_start.set_date(date.today())

        tk.Label(frame_dates, text="Fim:").pack(side=tk.LEFT)
        self.entry_end = DateEntry(frame_dates, width=12, background='darkblue',
                                   foreground='white', borderwidth=2, locale='pt_BR', date_pattern='dd/mm/yyyy')
        self.entry_end.pack(side=tk.LEFT, padx=5)
        self.entry_end.set_date(date.today() + timedelta(days=15))

        # Botões de Data Rápida
        frame_quick_dates = tk.Frame(root)
        frame_quick_dates.pack(pady=5)
        frame_quick_dates_2 = tk.Frame(root)
        frame_quick_dates_2.pack(pady=5)

        tk.Button(frame_quick_dates, text="Hoje", command=lambda: self.set_dates(0)).pack(side=tk.LEFT, padx=2)
        tk.Button(frame_quick_dates, text="Amanhã", command=lambda: self.set_dates(
            1, start_today=False)).pack(side=tk.LEFT, padx=2)
        tk.Button(frame_quick_dates, text="dps amanhã", command=lambda: self.set_dates(
            2, start_today=False)).pack(side=tk.LEFT, padx=2)
        tk.Button(frame_quick_dates_2, text="3 Dias", command=lambda: self.set_dates(3)).pack(side=tk.LEFT, padx=2)
        tk.Button(frame_quick_dates_2, text="7 Dias", command=lambda: self.set_dates(7)).pack(side=tk.LEFT, padx=2)
        tk.Button(frame_quick_dates_2, text="15 Dias", command=lambda: self.set_dates(15)).pack(side=tk.LEFT, padx=2)

        self.var_merge_days = tk.BooleanVar()
        self.chk_merge = tk.Checkbutton(
            root, text="Fundir dias em um único arquivo (Feriados)", variable=self.var_merge_days)
        self.chk_merge.pack(pady=5)

        self.progress = ttk.Progressbar(root, orient="horizontal", length=300, mode="indeterminate")
        self.progress.pack(pady=10)

        frame_actions = tk.Frame(root)
        frame_actions.pack(pady=20)

        self.btn_start = tk.Button(frame_actions, text="Iniciar Extração", command=self.start_thread,
                                   height=2, width=15, bg="#4CAF50", fg="black")
        self.btn_start.pack(side=tk.LEFT, padx=10)

        self.btn_cancel = tk.Button(frame_actions, text="Cancelar", command=self.cancel_process,
                                    height=2, width=15, bg="#f44336", fg="black", state="disabled")
        self.btn_cancel.pack(side=tk.LEFT, padx=10)

    def set_dates(self, days, start_today=True):
        today = date.today()
        if start_today:
            self.entry_start.set_date(today)
            self.entry_end.set_date(today + timedelta(days=days))
        else:
            # Starts at "today + days" and ends at "today + days"
            target_date = today + timedelta(days=days)
            self.entry_start.set_date(target_date)
            self.entry_end.set_date(target_date)

    def update_status(self, text):
        self.lbl_status.config(text=text)
        self.root.update_idletasks()

    def start_thread(self):
        self.stop_event.clear()
        self.btn_start.config(state="disabled")
        self.btn_cancel.config(state="normal")
        self.progress.start(10)
        threading.Thread(target=self.run_pipeline, daemon=True).start()

    def cancel_process(self):
        if not self.stop_event.is_set():
            self.update_status("Cancelando... Aguarde.")
            self.stop_event.set()

    def run_pipeline(self):
        import shutil
        try:
            import email_alert
        except ImportError:
            email_alert = None

        try:
            dt_ini = self.entry_start.get()
            dt_fim = self.entry_end.get()

            self.update_status("Verificando ambiente...")
            check_drive_access()

            # Pasta necessária para o report de erros do email
            if not os.path.exists(TEMP_DIR):
                os.makedirs(TEMP_DIR)

            # Etapa A: Buscar na API
            if self.stop_event.is_set():
                return
            df_xfin = fetch_xfin_data_api(self.update_status, dt_ini, dt_fim, self.stop_event)

            if self.stop_event.is_set():
                self.finish("Processo cancelado pelo usuário.")
                return

            # Etapa B: Processamento
            df_merged, missing, dt_start, dt_end, cols_map = process_data(
                df_xfin, self.update_status, self.stop_event, dt_ini_ui=dt_ini, dt_fim_ui=dt_fim)

            if self.stop_event.is_set():
                self.finish("Processo cancelado pelo usuário.")
                return

            if df_merged is None or df_merged.empty:
                self.finish("Nenhum pagamento encontrado para o período.")
                return

            # Etapa C: Salvar Arquivos
            self.update_status("Gerando planilha Excel...")
            print("Gerando planilha Excel...")

            # Nome do mês em PT-BR (simples)
            months = ["Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
                      "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]

            generated_files = []

            # Agrupar por Data de Arquivo (juntando FDS na Segunda)
            col_venc = cols_map[1]
            df_merged['File_Date'] = df_merged[col_venc].apply(get_file_date)

            print("Iniciando geração do arquivo único por data...")

            merge_days = self.var_merge_days.get()

            if merge_days:
                # Lógica de Fusão: Pega a data mais distante (max) para a pasta
                max_file_date = df_merged['File_Date'].max()
                min_file_date = df_merged['File_Date'].min()

                year = max_file_date.strftime("%Y")
                month_num = max_file_date.month
                month_name = months[month_num-1]
                folder_month = f"{month_num}. {month_name.upper()}"
                day_folder = max_file_date.strftime("%d-%m-%y")

                current_base_path = os.path.join(DRIVE_PATH, "CONTAS A PAGAR", year, folder_month, day_folder)
                if not os.path.exists(current_base_path):
                    os.makedirs(current_base_path)

                date_str_initial = min_file_date.strftime('%d_%m_%Y')
                date_str_final = max_file_date.strftime('%d_%m_%Y')

                if date_str_initial == date_str_final:
                    fname = f"Contas_A_Pagar-{date_str_final}.xlsx"
                else:
                    fname = f"Contas_A_Pagar-{date_str_initial}-{date_str_final}.xlsx"

                print(f"Processando (Fundido): {date_str_initial} a {date_str_final} - todas as filiais")
                full_path = os.path.join(current_base_path, fname)
                create_excel(df_merged, full_path, cols_map)
                generated_files.append(fname)
            else:
                # Um arquivo por data, com abas por filial
                for file_date, df_date in df_merged.groupby('File_Date'):
                    if self.stop_event.is_set():
                        self.finish("Processo cancelado pelo usuário.")
                        return

                    year = file_date.strftime("%Y")
                    month_num = file_date.month
                    month_name = months[month_num-1]
                    folder_month = f"{month_num}. {month_name.upper()}"
                    day_folder = file_date.strftime("%d-%m-%y")

                    current_base_path = os.path.join(DRIVE_PATH, "CONTAS A PAGAR", year, folder_month, day_folder)

                    if not os.path.exists(current_base_path):
                        os.makedirs(current_base_path)

                    date_str = file_date.strftime('%d_%m_%Y')
                    fname = f"Contas_A_Pagar-{date_str}.xlsx"

                    print(f"Processando: Data {date_str} - todas as filiais ({len(df_date)} registros)")
                    full_path = os.path.join(current_base_path, fname)
                    create_excel(df_date, full_path, cols_map)
                    generated_files.append(fname)

            print("Arquivo(s) gerado(s) com sucesso.")

            # Alerta de Faltantes
            if len(missing) > 0 and email_alert:
                if self.stop_event.is_set():
                    self.finish("Processo cancelado pelo usuário.")
                    return

                self.update_status("Enviando alerta de fornecedores...")
                # Cria um CSV temporário com os faltantes para anexar
                missing_df = pd.DataFrame(missing, columns=['Fornecedor'])
                missing_csv = os.path.join(TEMP_DIR, "falta_cadastrar.csv")
                missing_df.to_csv(missing_csv, index=False)
                email_alert.enviar_email_erro(missing_csv, len(missing), True)

            self.finish(f"Sucesso!\nGerado(s): {len(generated_files)} arquivo(s)\nSalvos em {current_base_path}.")

        except Exception as e:
            self.finish(f"Erro: {str(e)}", error=True)
        finally:
            # Limpeza
            if os.path.exists(TEMP_DIR):
                try:
                    shutil.rmtree(TEMP_DIR)
                except:
                    pass

    def finish(self, message, error=False):
        self.progress.stop()
        self.btn_start.config(state="normal")
        self.btn_cancel.config(state="disabled")
        self.update_status(message)
        if error:
            messagebox.showerror("Erro", message)
        else:
            messagebox.showinfo("Concluído", message)


if __name__ == "__main__":
    root = tk.Tk()
    app = PaymentBotApp(root)
    root.mainloop()
