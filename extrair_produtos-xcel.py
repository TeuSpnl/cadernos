import os
import firebirdsql
from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import Font

# Carregar variáveis de ambiente (mesmo .env do dynamo.py)
load_dotenv()

def get_firebird_connection():
    # Ajustar com os parâmetros corretos do Firebird, inclusive charset
    return firebirdsql.connect(
        host=os.getenv('HOST'),
        port=int(os.getenv('PORT', '3050')),
        database=os.getenv('DB_PATH'),
        user=os.getenv('APP_USER'),
        password=os.getenv('PASSWORD'),
        role=os.getenv('ROLE'),
        auth_plugin_name=os.getenv('AUTH'),
        wire_crypt=False,
        charset='ISO8859_1'
    )

def formatar_valor(valor):
    if valor is None:
        return "0,00"
    return f"{float(valor):.2f}".replace('.', ',')

def formatar_qtd(valor):
    if valor is None:
        return "0"
    return str(valor).replace('.', ',')

def processar_pedido(conn, numero):
    cursor = conn.cursor()
    # Cabeçalho: Data do Pedido
    cursor.execute("SELECT DATA FROM PEDIDOVENDA WHERE CDPEDIDOVENDA = ?", (numero,))
    row = cursor.fetchone()
    if not row:
        return None
    data_doc = row[0]

    # Itens: Baseado no dynamo.py (ITENSPEDIDOVENDA)
    # dynamo.py usa: NUMORIGINAL, DESCRICAO, QUANTIDADE, VALORUNITARIOCDESC
    cursor.execute("""
        SELECT NUMORIGINAL, DESCRICAO, QUANTIDADE, VALORUNITARIOCDESC
        FROM ITENSPEDIDOVENDA 
        WHERE CDPEDIDOVENDA = ?
    """, (numero,))
    
    itens = []
    for r in cursor.fetchall():
        num_orig, desc, qtd, val_unit = r
        # Calculando total pois dynamo.py não extrai VALORTOTAL explicitamente na query de itens
        val_total = (float(qtd) if qtd else 0) * (float(val_unit) if val_unit else 0)
        itens.append([num_orig, desc, qtd, val_unit, val_total])
        
    return data_doc, itens

def processar_nf(conn, numero):
    cursor = conn.cursor()
    # Cabeçalho: Buscar CDNOTASAIDA pelo NUMNOTA
    # Ordena por DATAEMISSAO DESC para pegar a mais recente caso haja duplicidade de número
    cursor.execute("SELECT FIRST 1 CDNOTASAIDA, DTSAIDAENTRADA FROM NOTASAIDA WHERE NUMNOTA = ? ORDER BY DTSAIDAENTRADA DESC", (numero,))
    row = cursor.fetchone()
    if not row:
        return None
    cd_nota, data_doc = row

    # Itens: ITENSNOTASAIDA
    # Colunas solicitadas: NUMORIGINAL, DESCRICAO, QUANT, VALORUNIT, VALORTOTAL
    cursor.execute("""
        SELECT NUMORIGINAL, DESCRICAO, QUANT, VALORUNIT, VALORTOTAL
        FROM ITENSNOTASAIDA 
        WHERE CDNOTASAIDA = ?
    """, (cd_nota,))
    
    itens = []
    for r in cursor.fetchall():
        # r já vem na ordem: [Original, Desc, Qtd, Unit, Total]
        itens.append(list(r))
        
    return data_doc, itens

def processar_os(conn, numero):
    cursor = conn.cursor()
    # Cabeçalho: ORDEMSERVICO (Data de Abertura)
    cursor.execute("SELECT DATA FROM ORDEMSERVICO WHERE CDORDEMSERVICO = ?", (numero,))
    row = cursor.fetchone()
    if not row:
        return None
    data_doc = row[0]

    # Itens: ITENSORDEMSERVICO
    # Colunas solicitadas: QUANTIDADE, CDSERVICO, DESCRICAO, VALORUNITARIO, VALORTOTAL
    cursor.execute("""
        SELECT CDSERVICO, DESCRICAO, QUANTIDADE, VALORUNITARIO, VALORTOTAL
        FROM ITENSORDEMSERVICO 
        WHERE CDORDEMSERVICO = ?
    """, (numero,))
    
    itens = []
    for r in cursor.fetchall():
        cd_servico, desc, qtd, val_unit, val_total = r
        # Padronizando ordem para saída: [Codigo, Descricao, Qtd, Unit, Total]
        itens.append([cd_servico, desc, qtd, val_unit, val_total])
        
    return data_doc, itens

def main():
    input_file = "lista_documentos.txt"
    output_file = "relatorio_produtos_servicos.xlsx"

    if not os.path.exists(input_file):
        print(f"ERRO: Arquivo '{input_file}' não encontrado.")
        print("Crie um arquivo de texto com uma linha por documento no formato: TIPO NUMERO")
        print("Exemplos:\nPEDIDO 12345\nNF 555\nOS 9090")
        return

    try:
        conn = get_firebird_connection()
    except Exception as e:
        print(f"Erro ao conectar no banco de dados: {e}")
        return

    print(f"Lendo {input_file} e gerando {output_file}...")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Relatorio"

    # Cabeçalho
    headers = ["Documento", "Data", "Código/Original", "Descrição", "Quantidade", "Valor Unit.", "Valor Total"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    row_idx = 2

    with open(input_file, 'r', encoding='utf-8') as f_in:

        for line in f_in:
            line = line.strip()
            if not line: continue
            
            parts = line.split()
            if len(parts) < 2:
                print(f"Linha ignorada (formato inválido): {line}")
                continue
            
            tipo = parts[0].upper()
            numero = parts[1]
            
            info = None
            try:
                if 'PEDIDO' in tipo:
                    info = processar_pedido(conn, numero)
                elif 'NF' in tipo or 'NOTA' in tipo:
                    info = processar_nf(conn, numero)
                elif 'OS' in tipo or 'ORDEM' in tipo:
                    info = processar_os(conn, numero)
            except Exception as e:
                print(f"Erro ao processar {tipo} {numero}: {e}")
                continue

            if info:
                data_doc, itens = info
                data_str = data_doc.strftime("%d/%m/%Y") if data_doc else ""
                
                # Linha 1: Documento e Data (Col 1 e 2 preenchidas)
                cell_doc = ws.cell(row=row_idx, column=1, value=f"{tipo} {numero}")
                cell_doc.font = Font(bold=True)
                
                cell_data = ws.cell(row=row_idx, column=2, value=data_str)
                cell_data.font = Font(bold=True)
                
                row_idx += 1
                
                # Linhas seguintes: Itens (Col 1 vazia para efeito hierárquico)
                for item in itens:
                    # item = [codigo, descricao, qtd, unit, total]
                    ws.cell(row=row_idx, column=3, value=item[0])
                    ws.cell(row=row_idx, column=4, value=item[1])
                    ws.cell(row=row_idx, column=5, value=formatar_qtd(item[2]))
                    ws.cell(row=row_idx, column=6, value=formatar_valor(item[3]))
                    ws.cell(row=row_idx, column=7, value=formatar_valor(item[4]))
                    row_idx += 1
                print(f"OK: {tipo} {numero}")
            else:
                print(f"NÃO ENCONTRADO: {tipo} {numero}")
                cell_nf = ws.cell(row=row_idx, column=1, value=f"{tipo} {numero} (NÃO ENCONTRADO)")
                cell_nf.font = Font(bold=True)
                row_idx += 1

    wb.save(output_file)
    conn.close()
    print("Processamento concluído.")

if __name__ == "__main__":
    main()
