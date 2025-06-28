import os
import datetime
import firebirdsql
import pandas as pd
import tkinter as tk
from tkinter import messagebox
from tkcalendar import DateEntry
from datetime import datetime, timedelta
from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo, TableColumn
from openpyxl.worksheet.page import PageMargins

# Carregar variáveis de ambiente
load_dotenv()

#########################
# CONFIGURAÇÕES GERAIS  #
#########################

# Cores de preenchimento e fonte para a tabela GERAL e TÍTULOS DE TABELAS ESPECÍFICAS
COLORS = {
    1: {'fill': None, 'font': None, 'delete_row': True},  # Linha será deletada
    2: {'fill': '9C0202', 'font': 'FFFFFF'},
    3: {'fill': '0F9ED5', 'font': '000000'},  # Banco do Brasil
    4: {'fill': 'FFFFFF', 'font': '000000'},
    5: {'fill': '3C7D22', 'font': '000000'},
    6: {'fill': 'FE9250', 'font': '000000'},  # Pix
    7: {'fill': 'DE4A4A', 'font': '000000'},
    8: {'fill': 'FFFF00', 'font': '000000'}  # Amarelo para DDA/NULL
}


def get_firebird_connection():
    # Ajustar com os parâmetros corretos do Firebird, inclusive charset
    return firebirdsql.connect(
        host=os.getenv('HOST'),
        port=int(os.getenv('PORT', '3050')),
        database=os.getenv('DB_PATH'),
        user=os.getenv('APP_USER'),
        password=os.getenv('PASSWORD'),
        role=os.getenv('ROLE'),
        auth_plugin_name=os.getenv('AUTH'),
        wire_crypt=False,
        charset='ISO8859_1'
    )


def fetch_data(start_date, end_date):
    """
    start_date e end_date devem ser strings no formato 'YYYY-MM-DD'.
    Retorna um DataFrame do pandas com os dados solicitados.
    """
    try:
        conn = get_firebird_connection()
        cursor = conn.cursor()

        # Monta o WHERE com as datas
        where_dates = f"""
            p.DTVENCIMENTO BETWEEN '{start_date}' AND '{end_date}'
        """

        # Consulta SQL para pegar as informações de APAGAR e CONTA_CREDITO
        # Adicionei COALESCE para tratar NUMCONTACRED nulo
        sql_query = f"""
            SELECT
                COALESCE(p.NUMCONTACRED, 8) AS NUMCONTACRED,
                cc.NOME AS NOME_CONTA_CREDITO,
                p.NUMDOCUMENTO,
                p.NOMEFORNECEDOR,
                p.VALOR,
                p.DESCRICAO
            FROM
                APAGAR p
            LEFT JOIN
                CONTA_CREDITO cc ON p.NUMCONTACRED = cc.NUMCONTACRED
            WHERE
                {where_dates}
            ORDER BY
                p.DTVENCIMENTO;
        """
        cursor.execute(sql_query)
        rows = cursor.fetchall()

        # Obter nomes das colunas da descrição do cursor
        columns = [desc[0] for desc in cursor.description]

        df = pd.DataFrame(rows, columns=columns)

        # Lidar com o caso de NUMCONTACRED = NULL, que já é tratado pelo COALESCE para 8
        # Mas vamos garantir que o nome da conta para NUMCONTACRED=8 (originalmente NULL) seja "DDA"
        if 8 in df['NUMCONTACRED'].unique():
            # Verifique se 'DDA' já é o nome para 8
            if not df[df['NUMCONTACRED'] == 8]['NOME_CONTA_CREDITO'].eq('DDA').all():
                # Se não for, atualize onde NUMCONTACRED é 8 e NOME_CONTA_CREDITO é nulo ou diferente de DDA
                df.loc[(df['NUMCONTACRED'] == 8) & (df['NOME_CONTA_CREDITO'].isnull()), 'NOME_CONTA_CREDITO'] = 'DDA'
                df.loc[(df['NUMCONTACRED'] == 8) & (df['NOME_CONTA_CREDITO'] != 'DDA'), 'NOME_CONTA_CREDITO'] = 'DDA'

        conn.close()
        return df

    except firebirdsql.OperationalError as e:
        messagebox.showerror("Erro de Conexão", f"Não foi possível conectar ao banco de dados: {e}")
        return pd.DataFrame()
    except Exception as e:
        messagebox.showerror("Erro na Consulta", f"Ocorreu um erro ao buscar os dados: {e}")
        return pd.DataFrame()


def generate_filename(start_date_str, end_date_str):
    """Gera o nome do arquivo Excel com base nas datas."""
    start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
    end_date = datetime.strptime(end_date_str, '%Y-%m-%d')

    if start_date == end_date:
        return f"arquivos/Contas_A_Pagar-{start_date.strftime('%d_%m_%Y')}.xlsx"
    else:
        return f"arquivos/Contas_A_Pagar-{start_date.strftime('%d_%m')}-{end_date.strftime('%d_%m_%Y')}.xlsx"


def apply_excel_formatting(ws, df_filtered, table_type="Geral", start_row_offset=0):
    """
    Aplica formatação específica, incluindo larguras de coluna, cores,
    e cria uma tabela Excel.
    """

    # Definir larguras de coluna
    ws.column_dimensions['A'].width = 20    # Largura (A): 20 - Nome da Conta de Crédito
    ws.column_dimensions['B'].width = 10    # Largura (B): 10 - Número do Documento
    ws.column_dimensions['C'].width = 42    # Largura (C): 42 - Nome do Fornecedor
    ws.column_dimensions['D'].width = 22    # Largura (D): 22 - CNPJ/Chave (para Pix)
    ws.column_dimensions['E'].width = 17    # Largura (E): 17 - Valor (para Pix)
    ws.column_dimensions['F'].width = 12.5  # Largura (F): 12.5 - Valor Total (DDA, Banco do Brasil, Outros)
    ws.column_dimensions['G'].width = 12.5  # Largura (G): 12.5 - Valor Total (Geral)

    ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.75, bottom=0.75, header=0.3, footer=0.3)

    value_column_header = "VALOR"

    if table_type == "DDA" or table_type == "Banco do Brasil" or table_type == "Outros":
        headers = ["Banco", "Nº DOC", "Nome", "OBSERVAÇÃO", "CNPJ", value_column_header]
        value_col_idx = 5  # F (0-indexed)
    elif table_type == "Pix":
        headers = ["Nome Recebedor", "Nº DOC", "Nome", "OBSERVAÇÃO", "Chave", value_column_header, "VALOR TOTAL"]
        value_col_idx = 5  # F (0-indexed)
        value_total_col_idx = 6  # G (0-indexed)
    else:  # Tabela Geral
        headers = ["RECURSO", "Nº DOC", "FORNECEDOR", "OBSERVAÇÃO", value_column_header]
        value_col_idx = 4  # E (0-indexed)

    current_row = start_row_offset

    # Escrever cabeçalhos da tabela
    for col_idx, header in enumerate(headers):
        cell = ws.cell(row=current_row, column=col_idx + 1, value=header)
        cell.font = Font(bold=True, color='000000')  # Títulos sempre pretos e negrito
        cell.border = Border(top=Side(style='thin'), bottom=Side(style='thin'),
                             left=Side(style='thin'), right=Side(style='thin'))

        numcontacred_for_header = 4  # Placeholder para cor geral, pode ser qualquer um que tenha fill definido
        fill_color_header = COLORS.get(
            numcontacred_for_header, {}).get(
            'fill', 'FFFFFF')  # Usa uma cor de preenchimento
        cell.fill = PatternFill(start_color=fill_color_header, end_color=fill_color_header, fill_type="solid")
    current_row += 1

    start_data_row = current_row  # Linha onde os dados começam para a tabela Excel

    # Escrever os dados e aplicar formatação de cor
    rows_to_delete = []
    original_row_idx = start_data_row  # Para rastrear o índice original da linha no Excel para deleção
    for index, row in df_filtered.iterrows():
        numcontacred = row['NUMCONTACRED']
        fill_color = COLORS.get(numcontacred, {}).get('fill')
        delete_row = COLORS.get(numcontacred, {}).get('delete_row', False)

        if delete_row:
            rows_to_delete.append(original_row_idx)
            original_row_idx += 1
            continue

        data_row = []

        # Usar a descrição do BD
        observacao = row['DESCRICAO'] if 'DESCRICAO' in row and row['DESCRICAO'] is not None else ""

        if table_type == "DDA" or table_type == "Banco do Brasil" or table_type == "Outros":
            data_row = ["", row['NUMDOCUMENTO'], row['NOMEFORNECEDOR'], observacao, "", row['VALOR']]
        elif table_type == "Pix":
            # Para Pix, VALOR_TOTAL já vem calculado no DataFrame
            data_row = [
                "",
                row['NUMDOCUMENTO'],
                row['NOMEFORNECEDOR'],
                observacao,
                "",
                row['VALOR'],
                row['VALOR_TOTAL']]
        else:  # Tabela Geral
            data_row = [row['NOME_CONTA_CREDITO'], row['NUMDOCUMENTO'], row['NOMEFORNECEDOR'], observacao, row['VALOR']]

        for col_idx, value in enumerate(data_row):
            cell = ws.cell(row=current_row, column=col_idx + 1, value=value)
            # Aplicar preenchimento apenas na tabela Geral
            if table_type == "Geral" and fill_color:
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            cell.border = Border(top=Side(style='thin'), bottom=Side(style='thin'),
                                 left=Side(style='thin'), right=Side(style='thin'))

            # Formatar colunas de valor como contábil
            if (table_type == "Geral" and col_idx == value_col_idx) or \
               ((table_type == "DDA" or table_type == "Banco do Brasil" or table_type == "Outros") and col_idx == value_col_idx) or \
               (table_type == "Pix" and (col_idx == value_col_idx or col_idx == value_total_col_idx)):
                cell.number_format = '_-R$* #,##0.00_-;-R$* #,##0.00_-;_-R$* \"-\"??_-;_-@_-'

        current_row += 1
        original_row_idx += 1

    # Deletar linhas marcadas
    # openpyxl deleta de baixo para cima para evitar problemas de índice
    for r_idx in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(r_idx, 1)
        # Ajustar start_data_row se as linhas deletadas forem antes do início da tabela
        if r_idx < start_data_row:
            start_data_row -= 1

     # Criar a tabela Excel
    if start_data_row < current_row:  # Verifica se há dados para formar a tabela
        # A linha de cabeçalho da tabela é (start_data_row - 1)
        table_ref = f"A{start_data_row-1}:{get_column_letter(len(headers))}{current_row-1}"
        tab_name = f"Table_{ws.title.replace(' ', '')}_{table_type.replace(' ', '')}_{start_row_offset-1}"  # Nome único

        # Criação explícita dos objetos TableColumn
        table_columns_list = []
        for i, header_name in enumerate(headers):
            tc = TableColumn(id=i+1, name=header_name)  # IDs de coluna começam de 1
            if table_type == "Pix" and i == value_total_col_idx:
                tc.totalsRowFunction = "sum"
                tc.totalsRowLabel = "Total:"
            elif table_type != "Pix" and i == value_col_idx:
                tc.totalsRowFunction = "sum"
                tc.totalsRowLabel = "Total:"
            table_columns_list.append(tc)

        tab = Table(displayName=tab_name, ref=table_ref, tableColumns=table_columns_list)  # Passa tableColumns aqui

        # Habilitar a linha de totais do Excel e definir estilo
        style = TableStyleInfo(name="TableStyleLight9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style
        tab.showTotalsRow = True

        ws.add_table(tab)

    # Lógica de mesclagem para a tabela Pix (fora da criação da Table)
    if table_type == "Pix":
        # Agrupar por fornecedor para mesclar as células de VALOR TOTAL
        current_supplier = None
        start_merge_row = -1

        # Iterar pelas linhas do DataFrame filtrado e mesclar na planilha
        row_in_excel = start_data_row  # A primeira linha de dados na planilha

        for idx, row in df_filtered.iterrows():
            if row['NOMEFORNECEDOR'] != current_supplier:
                # Se mudou de fornecedor, mesclar o bloco anterior (se houver)
                if start_merge_row != -1 and (row_in_excel - 1) > start_merge_row:
                    ws.merge_cells(start_row=start_merge_row, start_column=value_total_col_idx + 1,
                                   end_row=(row_in_excel - 1), end_column=value_total_col_idx + 1)
                    # Colocar o total no centro da célula mesclada
                    merged_cell = ws.cell(row=start_merge_row, column=value_total_col_idx + 1)
                    # O valor já está na primeira linha do bloco devido à ordenação
                    # Não precisamos pegar do df_filtered[df_filtered['NOMEFORNECEDOR'] == current_supplier]['VALOR_TOTAL'].iloc[0] novamente
                    merged_cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
                    merged_cell.font = Font(bold=True)
                    merged_cell.number_format = '_-R$* #,##0.00_-;-R$* #,##0.00_-;_-R$* \"-\"??_-;_-@_-'

                # Iniciar novo bloco de mesclagem
                current_supplier = row['NOMEFORNECEDOR']
                start_merge_row = row_in_excel

            # Escrever o valor na célula para a coluna "VALOR TOTAL" apenas na primeira linha do bloco
            if row_in_excel == start_merge_row:
                ws.cell(row=row_in_excel, column=value_total_col_idx + 1).value = row['VALOR_TOTAL']
            else:
                # Limpar células subsequentes para mesclar
                ws.cell(row=row_in_excel, column=value_total_col_idx + 1).value = ""

            # Aplicar formato numérico mesmo nas células vazias para a mesclagem funcionar bem
            ws.cell(row=row_in_excel, column=value_total_col_idx + 1).number_format = '_-R$* #,##0.00_-;-R$* #,##0.00_-;_-R$* \"-\"??_-;_-@_-'

            row_in_excel += 1  # Próxima linha na planilha

        # Mesclar o último bloco após o loop
        if start_merge_row != -1 and (row_in_excel - 1) >= start_merge_row:
            ws.merge_cells(start_row=start_merge_row, start_column=value_total_col_idx + 1,
                           end_row=(row_in_excel - 1), end_column=value_total_col_idx + 1)
            merged_cell = ws.cell(row=start_merge_row, column=value_total_col_idx + 1)
            # O valor já deve estar na primeira célula do bloco.
            merged_cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            merged_cell.font = Font(bold=True)
            merged_cell.number_format = '_-R$* #,##0.00_-;-R$* #,##0.00_-;_-R$* \"-\"??_-;_-@_-'

    current_row += 2  # Espaço entre tabelas (mantido para visualização clara entre as tabelas)

    return current_row


def fetch_data_and_generate_excel(start_date_str, end_date_str):
    df = fetch_data(start_date_str, end_date_str)

    if df.empty:
        messagebox.showinfo("Informação", "Nenhum dado encontrado para as datas selecionadas.")
        return

    try:
        # Deletar linhas onde NUMCONTACRED é 1
        df = df[df['NUMCONTACRED'] != 1].copy()  # Usar .copy() para evitar SettingWithCopyWarning

        # Criar um novo workbook e uma planilha
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Contas a Pagar"

        # Adicionar cabeçalho do arquivo no cabeçalho da página (header)
        filename_for_header = generate_filename(start_date_str, end_date_str).replace(".xlsx", "")
        ws.oddHeader.center.text = filename_for_header
        ws.oddHeader.center.font = "Arial,Bold"
        ws.oddHeader.center.size = 14

        current_row_offset = 1  # Começar a primeira tabela na linha 1

        # --- Tabela Geral ---
        df_geral = df.copy()
        # Ordenação da Tabela Geral
        df_geral = df_geral.sort_values(by=['NOME_CONTA_CREDITO', 'NOMEFORNECEDOR', 'VALOR'],
                                        ascending=[True, True, True])
        # Título da tabela geral, sem espaço extra
        current_row_offset = apply_excel_formatting(
            ws, df_geral, table_type="Geral", start_row_offset=current_row_offset)

        # --- Tabelas Específicas ---

        # DDA (NUMCONTACRED = 8 ou null)
        df_dda = df[(df['NUMCONTACRED'] == 8) | (df['NUMCONTACRED'].isnull())].copy()
        df_dda = df_dda.sort_values(by='VALOR', ascending=True)
        if not df_dda.empty:
            # Título da tabela DDA
            ws.cell(row=current_row_offset, column=1, value="DDA").font = Font(bold=True, size=12, color='000000')
            ws.cell(row=current_row_offset, column=1).fill = PatternFill(
                start_color='FFFF00', end_color='FFFF00', fill_type="solid")
            current_row_offset += 1  # Sem espaço extra
            current_row_offset = apply_excel_formatting(
                ws, df_dda, table_type="DDA", start_row_offset=current_row_offset)

        # Pix (NUMCONTACRED = 6)
        df_pix = df[df['NUMCONTACRED'] == 6].copy()
        # Calcular VALOR TOTAL para Pix antes de ordenar
        df_pix['VALOR_TOTAL'] = df_pix.groupby('NOMEFORNECEDOR')['VALOR'].transform('sum')
        # Ordenar Pix por VALOR_TOTAL (do menor para o maior)
        df_pix = df_pix.sort_values(by=['VALOR_TOTAL', 'NOMEFORNECEDOR', 'VALOR'], ascending=[True, True, True])
        if not df_pix.empty:
            # Título da tabela Pix
            ws.cell(row=current_row_offset, column=1, value="PIX").font = Font(bold=True, size=12, color='000000')
            ws.cell(row=current_row_offset, column=1).fill = PatternFill(
                start_color='FE9250', end_color='FE9250', fill_type="solid")
            current_row_offset += 1  # Sem espaço extra
            current_row_offset = apply_excel_formatting(
                ws, df_pix, table_type="Pix", start_row_offset=current_row_offset)

        # Banco do Brasil (NUMCONTACRED = 3)
        df_bb = df[df['NUMCONTACRED'] == 3].copy()
        df_bb = df_bb.sort_values(by='VALOR', ascending=True)
        if not df_bb.empty:
            # Título da tabela Banco do Brasil
            ws.cell(
                row=current_row_offset, column=1, value="BANCO DO BRASIL").font = Font(
                bold=True, size=12, color='000000')
            ws.cell(row=current_row_offset, column=1).fill = PatternFill(
                start_color='0F9ED5', end_color='0F9ED5', fill_type="solid")
            current_row_offset += 1  # Sem espaço extra
            current_row_offset = apply_excel_formatting(
                ws, df_bb, table_type="Banco do Brasil", start_row_offset=current_row_offset)

        # Demais Bancos
        # Obter os NUMCONTACREDs que já foram processados (1, 3, 6, 8)
        processed_numcontacreds = [1, 3, 6, 8]
        df_outros_bancos = df[~df['NUMCONTACRED'].isin(processed_numcontacreds)].copy()

        # Agrupar pelos demais NUMCONTACRED e gerar tabelas separadas
        for num_conta_cred_id in sorted(df_outros_bancos['NUMCONTACRED'].unique()):
            df_banco = df_outros_bancos[df_outros_bancos['NUMCONTACRED'] == num_conta_cred_id].copy()
            df_banco = df_banco.sort_values(by='VALOR', ascending=True)

            if not df_banco.empty:
                # Definir o preenchimento e fonte para o título da tabela do banco
                font_color = COLORS.get(num_conta_cred_id, {}).get('font')

                # Pegar o nome da conta para o título
                nome_conta = df_banco['NOME_CONTA_CREDITO'].iloc[0]
                # Pegar a cor de preenchimento para este NUMCONTACRED
                fill_color = COLORS.get(num_conta_cred_id, {}).get('fill', 'FFFFFF')  # Padrão branco se não definido

                # Título da tabela do banco
                ws.cell(row=current_row_offset, column=1, value=nome_conta.upper()
                        ).font = Font(bold=True, size=12, color=font_color)
                ws.cell(row=current_row_offset, column=1).fill = PatternFill(
                    start_color=fill_color, end_color=fill_color, fill_type="solid")
                current_row_offset += 1  # Sem espaço extra
                current_row_offset = apply_excel_formatting(
                    ws, df_banco, table_type="Outros", start_row_offset=current_row_offset)

    except Exception as e:
        messagebox.showerror(
            "Erro ao Gerar Excel", f"Ocorreu um erro ao gerar o arquivo Excel: {e}\n {e.__str__}\n {e.__class__}\n {e.__dict__}")
        return

    # Salvar o arquivo Excel
    try:
        file_path = os.path.join(os.getcwd(), generate_filename(start_date_str, end_date_str))
        wb.save(file_path)
        messagebox.showinfo(
            "Sucesso", f"Dados exportados com sucesso para: {generate_filename(start_date_str, end_date_str)}")
    except Exception as e:
        messagebox.showerror("Erro ao Salvar", f"Não foi possível salvar o arquivo Excel: {e}")


# --- Funções para a interface gráfica ---
def fill_today():
    today = datetime.now()
    start_date_var.set(today.strftime('%d/%m/%Y'))
    end_date_var.set(today.strftime('%d/%m/%Y'))


def fill_tomorrow():
    today = datetime.now()
    tomorrow = today + timedelta(days=1)
    start_date_var.set(tomorrow.strftime('%d/%m/%Y'))
    end_date_var.set(tomorrow.strftime('%d/%m/%Y'))


def fill_after_tomorrow():
    today = datetime.now()
    after_tomorrow = today + timedelta(days=2)
    start_date_var.set(after_tomorrow.strftime('%d/%m/%Y'))
    end_date_var.set(after_tomorrow.strftime('%d/%m/%Y'))


def on_generate_report():
    try:
        start_date_str_display = start_date_var.get()
        end_date_str_display = end_date_var.get()

        start_date_obj = datetime.strptime(start_date_str_display, '%d/%m/%Y')
        end_date_obj = datetime.strptime(end_date_str_display, '%d/%m/%Y')

        # Formato YYYY-MM-DD para a função fetch_data
        start_date_db_format = start_date_obj.strftime('%Y-%m-%d')
        end_date_db_format = end_date_obj.strftime('%Y-%m-%d')

        fetch_data_and_generate_excel(start_date_db_format, end_date_db_format)
    except ValueError:
        messagebox.showerror("Erro de Formato", "Por favor, insira as datas no formato DD/MM/YYYY.")
    except Exception as e:
        messagebox.showerror("Erro", f"Ocorreu um erro inesperado: {e}")


# Configuração da interface Tkinter
root = tk.Tk()
root.title("Gerador de Relatório de Contas a Pagar")

# Frame para entradas de data
frame_inputs = tk.Frame(root)
frame_inputs.pack(padx=10, pady=10)

# Variáveis para as datas
start_date_var = tk.StringVar()
end_date_var = tk.StringVar()

# Data inicial
tk.Label(frame_inputs, text="Data Inicial (DD/MM/YYYY):").grid(row=0, column=0, sticky=tk.W, padx=5, pady=5)
start_date_entry = DateEntry(
    frame_inputs,
    textvariable=start_date_var,
    date_pattern='dd/MM/yyyy',
    locale='pt_BR'
)
start_date_entry.grid(row=0, column=1, padx=5, pady=5)

# Data final
tk.Label(frame_inputs, text="Data Final (DD/MM/YYYY):").grid(row=1, column=0, sticky=tk.W, padx=5, pady=5)
end_date_entry = DateEntry(
    frame_inputs,
    textvariable=end_date_var,
    date_pattern='dd/MM/yyyy',
    locale='pt_BR'
)
end_date_entry.grid(row=1, column=1, padx=5, pady=5)

# Frame para botões de período
frame_buttons = tk.Frame(root)
frame_buttons.pack(padx=10, pady=10, fill=tk.X)

btn_tomorrow = tk.Button(frame_buttons, text="Amanhã", command=fill_tomorrow)
btn_tomorrow.pack(side=tk.LEFT, padx=5)

btn_today = tk.Button(frame_buttons, text="Hoje", command=fill_today)
btn_today.pack(side=tk.LEFT, padx=5)

btn_after_tomorrow = tk.Button(frame_buttons, text="Depois de amanhã", command=fill_after_tomorrow)
btn_after_tomorrow.pack(side=tk.LEFT, padx=5)

# Botão para gerar relatório
btn_generate = tk.Button(root, text="Gerar Relatório", command=on_generate_report)
btn_generate.pack(pady=10)

root.mainloop()
