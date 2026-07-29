"""Exportacao Excel do painel admin (backup + producao de camisetas)."""

from __future__ import annotations

from datetime import datetime
from io import BytesIO

from django.db.models import Count, Q, Sum
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import Camiseta, ItemPedido, Pedido, Usuario


def _estilo_cabecalho(ws, colunas: int) -> None:
    fill = PatternFill("solid", fgColor="2F4632")
    font = Font(bold=True, color="FFF8F0")
    for col in range(1, colunas + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _ajustar_larguras(ws, larguras: list[int]) -> None:
    for idx, width in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def _itens_ativos():
    return ItemPedido.objects.exclude(
        pedido__status_pagamento=Pedido.StatusPagamento.CANCELADO
    )


def _aba_producao(wb: Workbook) -> None:
    """Planilha principal para quem vai confeccionar: modelo + tamanho + qtd."""
    ws = wb.active
    ws.title = "Producao"
    ws.append(["Modelo", "Categoria", "Tamanho", "Quantidade"])
    _estilo_cabecalho(ws, 4)

    labels = dict(Camiseta.TAMANHOS_TODOS)
    ordem = {v: i for i, (v, _) in enumerate(Camiseta.TAMANHOS_TODOS)}
    camisetas = {
        c.id: c
        for c in Camiseta.objects.filter(ativo=True).order_by(
            "categoria", "-destaque", "nome"
        )
    }

    rows = list(
        _itens_ativos()
        .values("camiseta_id", "nome_camiseta", "tamanho")
        .annotate(qtd=Sum("quantidade"))
    )
    rows.sort(
        key=lambda r: (
            camisetas[r["camiseta_id"]].categoria
            if r["camiseta_id"] in camisetas
            else "z",
            r["nome_camiseta"],
            ordem.get(r["tamanho"], 999),
        )
    )

    for row in rows:
        camiseta = camisetas.get(row["camiseta_id"])
        categoria = (
            camiseta.get_categoria_display() if camiseta else ""
        )
        ws.append(
            [
                row["nome_camiseta"],
                categoria,
                labels.get(row["tamanho"], row["tamanho"]),
                row["qtd"] or 0,
            ]
        )

    # Totais por modelo (linhas com qtd 0 tambem, para conferencia)
    ws2_start = len(rows) + 3
    ws.cell(row=ws2_start, column=1, value="Totais por modelo").font = Font(
        bold=True
    )
    for camiseta in camisetas.values():
        total = (
            _itens_ativos()
            .filter(camiseta_id=camiseta.id)
            .aggregate(t=Sum("quantidade"))["t"]
            or 0
        )
        ws2_start += 1
        ws.cell(row=ws2_start, column=1, value=camiseta.nome)
        ws.cell(row=ws2_start, column=2, value=camiseta.get_categoria_display())
        ws.cell(row=ws2_start, column=4, value=total)

    _ajustar_larguras(ws, [32, 12, 16, 12])


def _aba_visao_geral(wb: Workbook) -> None:
    ws = wb.create_sheet("Visao Geral")
    ws.append(["Modelo", "Categoria", "Qtd total", "Detalhe por tamanho"])
    _estilo_cabecalho(ws, 4)

    labels = dict(Camiseta.TAMANHOS_TODOS)
    ordem = {v: i for i, (v, _) in enumerate(Camiseta.TAMANHOS_TODOS)}
    por_tamanho: dict[int, list[str]] = {}
    for row in (
        _itens_ativos()
        .values("camiseta_id", "tamanho")
        .annotate(qtd=Sum("quantidade"))
    ):
        por_tamanho.setdefault(row["camiseta_id"], []).append(
            (row["tamanho"], row["qtd"] or 0)
        )

    for camiseta in Camiseta.objects.filter(ativo=True).order_by(
        "categoria", "-destaque", "nome"
    ):
        pares = por_tamanho.get(camiseta.id, [])
        pares.sort(key=lambda p: ordem.get(p[0], 999))
        detalhe = ", ".join(
            f"{qtd}x {labels.get(tam, tam)}" for tam, qtd in pares
        ) or "-"
        total = sum(q for _, q in pares)
        ws.append(
            [
                camiseta.nome,
                camiseta.get_categoria_display(),
                total,
                detalhe,
            ]
        )

    _ajustar_larguras(ws, [32, 12, 12, 48])


def _aba_por_igreja(wb: Workbook) -> None:
    ws = wb.create_sheet("Por Igreja")
    ws.append(
        [
            "Igreja",
            "Total camisetas",
            "Pedidos",
            "Total arrecadado",
            "Detalhe (modelo + tamanho)",
        ]
    )
    _estilo_cabecalho(ws, 5)

    labels = dict(Camiseta.TAMANHOS_TODOS)
    ordem = {v: i for i, (v, _) in enumerate(Camiseta.TAMANHOS_TODOS)}
    itens = _itens_ativos()

    totais = {
        (row["pedido__cliente__igreja"] or "Nao informada"): row["total"] or 0
        for row in itens.values("pedido__cliente__igreja").annotate(
            total=Sum("quantidade")
        )
    }
    pedidos_agg = {
        (row["cliente__igreja"] or "Nao informada"): row
        for row in Pedido.objects.exclude(
            status_pagamento=Pedido.StatusPagamento.CANCELADO
        )
        .values("cliente__igreja")
        .annotate(qtd_pedidos=Count("id"), valor_total=Sum("valor_total"))
    }
    detalhes: dict[str, list] = {}
    for row in itens.values(
        "pedido__cliente__igreja", "nome_camiseta", "tamanho"
    ).annotate(qtd=Sum("quantidade")):
        igreja = row["pedido__cliente__igreja"] or "Nao informada"
        detalhes.setdefault(igreja, []).append(
            (
                row["nome_camiseta"],
                row["tamanho"],
                row["qtd"] or 0,
            )
        )

    igrejas = sorted(totais.keys(), key=lambda i: totais[i], reverse=True)
    for igreja in igrejas:
        linhas = detalhes.get(igreja, [])
        linhas.sort(key=lambda x: (x[0], ordem.get(x[1], 999)))
        detalhe = "; ".join(
            f"{qtd}x {nome} ({labels.get(tam, tam)})"
            for nome, tam, qtd in linhas
        )
        agg = pedidos_agg.get(igreja, {})
        ws.append(
            [
                igreja,
                totais[igreja],
                agg.get("qtd_pedidos", 0),
                float(agg.get("valor_total") or 0),
                detalhe,
            ]
        )

    _ajustar_larguras(ws, [40, 16, 10, 16, 60])


def _aba_por_cliente(wb: Workbook) -> None:
    ws = wb.create_sheet("Por Cliente")
    ws.append(["Cliente", "Usuario", "Telefone", "Igreja", "Pedidos", "Total gasto"])
    _estilo_cabecalho(ws, 6)

    clientes = (
        Usuario.objects.annotate(
            total_gasto=Sum(
                "pedidos__valor_total",
                filter=~Q(
                    pedidos__status_pagamento=Pedido.StatusPagamento.CANCELADO
                ),
            ),
            qtd_pedidos=Count(
                "pedidos",
                filter=~Q(
                    pedidos__status_pagamento=Pedido.StatusPagamento.CANCELADO
                ),
            ),
        )
        .order_by("-total_gasto", "nome_completo")
    )
    for c in clientes:
        ws.append(
            [
                c.nome_completo,
                c.username,
                c.telefone or "",
                c.igreja or "",
                c.qtd_pedidos or 0,
                float(c.total_gasto or 0),
            ]
        )

    _ajustar_larguras(ws, [28, 16, 16, 40, 10, 14])


def _aba_pedidos(wb: Workbook) -> None:
    ws = wb.create_sheet("Pedidos")
    ws.append(
        [
            "Codigo",
            "Data",
            "Cliente",
            "Telefone",
            "Igreja",
            "Item",
            "Tamanho",
            "Qtd",
            "Preco unit.",
            "Subtotal",
            "Total pedido",
            "Pagamento",
            "Entrega",
            "Observacao",
        ]
    )
    _estilo_cabecalho(ws, 14)

    labels = dict(Camiseta.TAMANHOS_TODOS)
    pedidos = (
        Pedido.objects.select_related("cliente")
        .prefetch_related("itens")
        .order_by("-criado_em")
    )
    for pedido in pedidos:
        cliente = pedido.cliente
        base = [
            pedido.codigo_curto,
            pedido.criado_em.strftime("%d/%m/%Y %H:%M"),
            cliente.nome_completo,
            cliente.telefone or "",
            cliente.igreja or "",
        ]
        itens = list(pedido.itens.all())
        if not itens:
            ws.append(
                base
                + [
                    "",
                    "",
                    "",
                    "",
                    "",
                    float(pedido.valor_total),
                    pedido.get_status_pagamento_display(),
                    pedido.get_status_entrega_display(),
                    pedido.observacoes or "",
                ]
            )
            continue
        for item in itens:
            ws.append(
                base
                + [
                    item.nome_camiseta,
                    labels.get(item.tamanho, item.tamanho),
                    item.quantidade,
                    float(item.preco_unitario),
                    float(item.subtotal),
                    float(pedido.valor_total),
                    pedido.get_status_pagamento_display(),
                    pedido.get_status_entrega_display(),
                    pedido.observacoes or "",
                ]
            )

    _ajustar_larguras(
        ws, [10, 16, 24, 14, 32, 28, 12, 6, 12, 12, 12, 12, 12, 28]
    )


def gerar_excel_painel() -> BytesIO:
    """Monta o workbook completo e devolve um buffer pronto para download."""
    wb = Workbook()
    _aba_producao(wb)
    _aba_visao_geral(wb)
    _aba_por_igreja(wb)
    _aba_por_cliente(wb)
    _aba_pedidos(wb)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def nome_arquivo_excel() -> str:
    stamp = datetime.now().strftime("%Y%m%d_%H%M")
    return f"feira_camisetas_{stamp}.xlsx"
